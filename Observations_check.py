"""
Observations_check.py
=====================
Checks for observational CTD cast data from the MVN Saltwater Wedge field data
directory, processes measurements collected within the last 7 days, and writes
observed toe position and near-surface salinity front locations to the
MRSWAT_Data.xlsx spreadsheet.

Workflow:
  1. Scans the field data network directory for SWW_YYYYMMDD folders whose dates
     fall within the last 7 days from today.
  2. Parses all CTD cast CSV files inside qualifying folders, extracting GPS
     coordinates from the metadata and salinity/depth profiles from the
     measurement matrix.
  3. Converts each cast's Start latitude / longitude to a river mile using the
     nearest reference point in river_miles.csv (MISSISSIPPI-LO).
  4. For each observation day:
       - Toe position: upstream-most (highest) river mile where any depth 
         records salinity >= 9 ppt.
       - Near-surface 0.25 ppt front: upstream-most river mile where the
         near-surface water column (within 10 ft of surface) exceeds 0.25 ppt.
         If no such location exists, assumes 15 river miles downstream of toe.
  5. Writes results to the 'Observations' sheet in MRSWAT_Data.xlsx, skipping
     dates that already exist in the spreadsheet.
"""

import os
import sys
import csv
import math
import glob
import re
from datetime import datetime, timedelta
import openpyxl


# ============================================================================
#  CONFIGURATION
# ============================================================================

# Network path to CTD field data
FIELD_DATA_DIR = (
    r"\\mvd\mvn\Data_EDR\02 Stream Gaging Section"
    r"\Gaging Stations\Gages_RandomProjects\Saltwater Wedge\Field Data"
)

# Directory paths derived from script location
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
SIM_DIR = os.path.dirname(SCRIPT_DIR)       # simulation folder (parent of Code Storage)
FORECAST_DIR = os.path.dirname(SIM_DIR)      # parent of simulation folder
RIVER_MILES_CSV = os.path.join(SCRIPT_DIR, "river_miles.csv")
EXCEL_PATH = os.path.join(FORECAST_DIR, "MRSWAT_Data.xlsx")

# Metadata layout
METADATA_LINES = 27      # lines 1-27 are metadata
HEADER_LINE_INDEX = 28    # 0-based index of column-header row (line 29)
DATA_START_INDEX = 29     # 0-based index of first data row (line 30)

# Analysis parameters
LOOKBACK_DAYS = 7
TOE_SALINITY_PPT = 9.0
SURFACE_SALINITY_PPT = 0.25
SURFACE_DEPTH_FT = 10.0
DOWNSTREAM_OFFSET_MILES = 15.0


# ============================================================================
#  UTILITY FUNCTIONS
# ============================================================================

def haversine_distance(lat1, lon1, lat2, lon2):
    """Great-circle distance in statute miles between two WGS-84 points."""
    R = 3958.8  # Earth radius in miles
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlam = math.radians(lon2 - lon1)
    a = (math.sin(dphi / 2) ** 2 +
         math.cos(phi1) * math.cos(phi2) * math.sin(dlam / 2) ** 2)
    return 2 * R * math.atan2(math.sqrt(a), math.sqrt(1 - a))


def load_river_miles(csv_path):
    """
    Load MISSISSIPPI-LO river mile reference points from river_miles.csv.
    Returns a list of dicts sorted by ascending mile.
    """
    points = []
    with open(csv_path, "r") as fh:
        reader = csv.DictReader(fh)
        for row in reader:
            if row["RIVER_NAME"] == "MISSISSIPPI-LO":
                points.append({
                    "mile": float(row["MILE"]),
                    "lat":  float(row["LATITUDE1"]),
                    "lon":  float(row["LONGITUDE1"]),
                })
    points.sort(key=lambda p: p["mile"])
    return points


def find_nearest_river_mile(lat, lon, ref_points):
    """
    Return (river_mile, distance_miles) for the closest reference point.
    """
    best_mile = None
    best_dist = float("inf")
    for pt in ref_points:
        d = haversine_distance(lat, lon, pt["lat"], pt["lon"])
        if d < best_dist:
            best_dist = d
            best_mile = pt["mile"]
    return best_mile, best_dist


# ============================================================================
#  CTD FILE PARSER
# ============================================================================

def parse_ctd_csv(filepath):
    """
    Parse a single CTD cast CSV file.

    Returns a dict with:
      start_lat, start_lon  – GPS coordinates from metadata
      data                  – list of row dicts (numeric values keyed by header)
      headers               – list of column header strings
    Returns None if the file cannot be parsed.
    """
    with open(filepath, "r") as fh:
        lines = fh.readlines()

    if len(lines) < DATA_START_INDEX + 1:
        return None  # file too short to contain data

    # --- Parse metadata (lines 1-27) ---
    metadata = {}
    for line in lines[:METADATA_LINES]:
        line = line.strip()
        if line.startswith("% ") and "," in line:
            key, _, val = line[2:].partition(",")
            metadata[key.strip()] = val.strip()

    try:
        start_lat = float(metadata["Start latitude"])
        start_lon = float(metadata["Start longitude"])
    except (KeyError, ValueError):
        return None  # missing or bad coordinates

    # --- Parse column headers (line 29, index 28) ---
    headers = [h.strip() for h in lines[HEADER_LINE_INDEX].strip().split(",")]

    # --- Parse measurement rows (line 30+, index 29+) ---
    data_rows = []
    for line in lines[DATA_START_INDEX:]:
        line = line.strip()
        if not line:
            continue
        values = line.split(",")
        if len(values) != len(headers):
            continue
        row = {}
        for j, h in enumerate(headers):
            try:
                row[h] = float(values[j])
            except ValueError:
                row[h] = values[j]
        data_rows.append(row)

    return {
        "start_lat": start_lat,
        "start_lon": start_lon,
        "metadata":  metadata,
        "headers":   headers,
        "data":      data_rows,
        "filepath":  filepath,
    }


# ============================================================================
#  DATE-FOLDER DISCOVERY
# ============================================================================

def find_recent_date_folders(base_dir, today, lookback_days):
    """
    Scan year sub-folders for SWW_YYYYMMDD directories whose date falls within
    [today - lookback_days, today].  Returns a dict {date_str: folder_path}.
    """
    cutoff = today - timedelta(days=lookback_days)
    results = {}  # date_str (YYYYMMDD) -> full folder path

    # Determine which year folders to inspect
    years_to_check = set()
    d = cutoff
    while d <= today:
        years_to_check.add(str(d.year))
        d += timedelta(days=1)

    for year_str in sorted(years_to_check):
        year_dir = os.path.join(base_dir, year_str)
        if not os.path.isdir(year_dir):
            continue
        for entry in os.listdir(year_dir):
            match = re.match(r"^SWW_(\d{8})$", entry)
            if not match:
                continue
            date_str = match.group(1)
            try:
                folder_date = datetime.strptime(date_str, "%Y%m%d")
            except ValueError:
                continue
            if cutoff <= folder_date <= today:
                results[date_str] = os.path.join(year_dir, entry)
    return results


# ============================================================================
#  MAIN PROCESSING
# ============================================================================

def main():
    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    cutoff = today - timedelta(days=LOOKBACK_DAYS)

    print("=" * 78)
    print("  OBSERVATIONS CHECK  –  CTD Saltwater Wedge Field Data Processor")
    print("=" * 78)
    print(f"  Current date        : {today.strftime('%Y-%m-%d')}")
    print(f"  Lookback window     : {cutoff.strftime('%Y-%m-%d')} to {today.strftime('%Y-%m-%d')}")
    print(f"  Field data directory: {FIELD_DATA_DIR}")
    print(f"  River miles CSV     : {RIVER_MILES_CSV}")
    print(f"  Output Excel file   : {EXCEL_PATH}")
    print()

    # ------------------------------------------------------------------
    # 1. Verify paths
    # ------------------------------------------------------------------
    if not os.path.isdir(FIELD_DATA_DIR):
        print("  ERROR: Field data directory not found. Check network connectivity.")
        sys.exit(1)
    print("  [OK] Field data directory is accessible.")

    if not os.path.isfile(RIVER_MILES_CSV):
        print("  ERROR: river_miles.csv not found in Code Storage folder.")
        sys.exit(1)
    print("  [OK] river_miles.csv located.")

    if not os.path.isfile(EXCEL_PATH):
        print("  ERROR: MRSWAT_Data.xlsx not found at expected location.")
        sys.exit(1)
    print("  [OK] MRSWAT_Data.xlsx located.")
    print()

    # ------------------------------------------------------------------
    # 2. Load river mile reference points
    # ------------------------------------------------------------------
    print("  Loading river mile reference data …")
    ref_points = load_river_miles(RIVER_MILES_CSV)
    print(f"    Loaded {len(ref_points)} MISSISSIPPI-LO reference points")
    print(f"    Mile range: {ref_points[0]['mile']:.0f} – {ref_points[-1]['mile']:.0f}")
    print()

    # ------------------------------------------------------------------
    # 3. Discover date folders within the lookback window
    # ------------------------------------------------------------------
    print("  Scanning for observation folders within the lookback window …")
    date_folders = find_recent_date_folders(FIELD_DATA_DIR, today, LOOKBACK_DAYS)

    if not date_folders:
        print("    No CTD observation folders found within the last"
              f" {LOOKBACK_DAYS} days.")
        print("    Nothing to process. Exiting.")
        return

    for ds, path in sorted(date_folders.items()):
        print(f"    Found: {ds}  →  {path}")
    print()

    # ------------------------------------------------------------------
    # 4. Load existing observation dates from the Excel file
    # ------------------------------------------------------------------
    print("  Reading existing observations from MRSWAT_Data.xlsx …")
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["Observations"]

    existing_dates = set()
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        val = row[0]
        if val is not None:
            existing_dates.add(str(val).strip())
    print(f"    {len(existing_dates)} existing observation date(s) in spreadsheet.")
    print()

    # ------------------------------------------------------------------
    # 5. Process each qualifying date folder
    # ------------------------------------------------------------------
    rows_written = 0

    for date_str in sorted(date_folders.keys()):
        folder_path = date_folders[date_str]
        print("-" * 78)
        print(f"  Processing observation date: {date_str}")
        print(f"    Folder: {folder_path}")

        # Check for duplicate
        if date_str in existing_dates:
            print(f"    SKIPPED – data for {date_str} already exists in the spreadsheet.")
            continue

        # Collect CSV cast files (exclude Summary.csv and any non-cast files)
        csv_files = glob.glob(os.path.join(folder_path, "*.csv"))
        cast_files = [
            f for f in csv_files
            if not os.path.basename(f).lower().startswith("summary")
        ]
        print(f"    Found {len(cast_files)} CTD cast file(s).")

        if not cast_files:
            print("    WARNING – no cast files found; skipping this date.")
            continue

        # Parse each cast and compute river mile + salinity metrics
        cast_results = []  # list of dicts with river_mile, max_salinity, max_surface_salinity

        for cf in sorted(cast_files):
            fname = os.path.basename(cf)
            cast = parse_ctd_csv(cf)
            if cast is None:
                print(f"      {fname}: SKIPPED (could not parse)")
                continue

            # Find river mile
            rm, dist = find_nearest_river_mile(
                cast["start_lat"], cast["start_lon"], ref_points
            )
            if dist > 2.0:
                print(f"      {fname}: WARNING – nearest river mile is {dist:.2f} mi away "
                      f"(lat={cast['start_lat']:.6f}, lon={cast['start_lon']:.6f})")

            # Identify salinity/depth column names dynamically
            depth_col = None
            sal_col = None
            for h in cast["headers"]:
                hl = h.lower()
                if "depth" in hl and "foot" in hl:
                    depth_col = h
                if "salinity" in hl:
                    sal_col = h
            if depth_col is None or sal_col is None:
                print(f"      {fname}: SKIPPED – could not identify depth/salinity columns")
                continue

            # Compute max salinity at any depth
            max_sal = max(
                (row[sal_col] for row in cast["data"]
                 if isinstance(row.get(sal_col), (int, float))),
                default=0.0,
            )

            # Compute max salinity in the near-surface zone (depth <= 10 ft)
            surface_sals = [
                row[sal_col]
                for row in cast["data"]
                if isinstance(row.get(depth_col), (int, float))
                and isinstance(row.get(sal_col), (int, float))
                and row[depth_col] <= SURFACE_DEPTH_FT
            ]
            max_surf_sal = max(surface_sals) if surface_sals else 0.0

            cast_results.append({
                "file":             fname,
                "river_mile":       rm,
                "dist_to_ref":      dist,
                "max_salinity":     max_sal,
                "max_surf_sal":     max_surf_sal,
                "n_rows":           len(cast["data"]),
                "max_depth":        max(
                    (row[depth_col] for row in cast["data"]
                     if isinstance(row.get(depth_col), (int, float))),
                    default=0.0,
                ),
            })

            print(f"      {fname}: RM {rm:.1f}  |  max sal = {max_sal:.2f} ppt  "
                  f"|  surface sal = {max_surf_sal:.4f} ppt  |  max depth = "
                  f"{cast_results[-1]['max_depth']:.1f} ft  |  {len(cast['data'])} rows")

        if not cast_results:
            print("    WARNING – no usable cast data; skipping this date.")
            continue

        # ------------------------------------------------------------------
        # Compute toe position (upstream-most RM with salinity >= 9 ppt)
        # ------------------------------------------------------------------
        toe_candidates = [
            cr for cr in cast_results if cr["max_salinity"] >= TOE_SALINITY_PPT
        ]
        if toe_candidates:
            toe_rm = max(cr["river_mile"] for cr in toe_candidates)
            print(f"\n    Toe position (>= {TOE_SALINITY_PPT} ppt): "
                  f"River Mile {toe_rm:.1f}")
        else:
            print(f"\n    WARNING – no cast recorded salinity >= {TOE_SALINITY_PPT} ppt.")
            print("    Cannot determine toe position; skipping this date.")
            continue

        # ------------------------------------------------------------------
        # Compute 0.25 ppt near-surface front
        # ------------------------------------------------------------------
        surf_candidates = [
            cr for cr in cast_results
            if cr["max_surf_sal"] > SURFACE_SALINITY_PPT
        ]
        if surf_candidates:
            surf_rm = max(cr["river_mile"] for cr in surf_candidates)
            surf_method = "measured"
            print(f"    0.25-ppt surface front: River Mile {surf_rm:.1f}  (measured)")
        else:
            surf_rm = toe_rm - DOWNSTREAM_OFFSET_MILES
            surf_method = "estimated (no measured location >= 0.25 ppt near surface)"
            print(f"    0.25-ppt surface front: River Mile {surf_rm:.1f}  ({surf_method})")

        # ------------------------------------------------------------------
        # Write to Excel
        # ------------------------------------------------------------------
        next_row = ws.max_row + 1
        ws.cell(row=next_row, column=1, value=date_str)
        ws.cell(row=next_row, column=2, value=round(toe_rm, 2))
        ws.cell(row=next_row, column=3, value=round(surf_rm, 2))
        existing_dates.add(date_str)
        rows_written += 1
        print(f"    >> Written to Observations sheet row {next_row}: "
              f"Date={date_str}, Toe RM={toe_rm:.2f}, Surface 0.25ppt RM={surf_rm:.2f}")

    # ------------------------------------------------------------------
    # 6. Save workbook
    # ------------------------------------------------------------------
    print()
    print("=" * 78)
    if rows_written > 0:
        wb.save(EXCEL_PATH)
        print(f"  MRSWAT_Data.xlsx saved – {rows_written} new observation row(s) written.")
    else:
        print("  No new observations to write.")
    wb.close()

    print("=" * 78)
    print("  Observations check complete.")
    print("=" * 78)


if __name__ == "__main__":
    main()
